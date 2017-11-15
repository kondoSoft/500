from django.core import serializers
from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.http import Http404, HttpResponse, JsonResponse
from django.template import loader
from django.contrib.auth import authenticate, login
from django.conf import settings
from django.core.mail import send_mail
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook, load_workbook
from .forms import SignUpForm, PerfilForm, Pregunta_Rechazada_Form
from .models import Empresa, Cuestionario, Pregunta, Articulo, Catalogo_Preguntas, Respuestas, Sectores, Paises, Fuentes, Glosario, Entradas_Recientes, Perfil,Pregunta_Rechazada
from .models import Corte
from django.contrib.auth.models import User, Group
from django.urls import reverse_lazy


from django.views.generic import ListView, DetailView

from django.views.generic.edit import CreateView, UpdateView, DeleteView

import json
from django.core.serializers.json import DjangoJSONEncoder

from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.mixins import PermissionRequiredMixin


#### CRUD PREGUNTA RECHAZADA

class CreatePreguntaRechazada(CreateView):
    model = Pregunta_Rechazada
    fields = ['motivo', 'respuestaPersonalizada', 'comentarios',]
    template_name = 'pregunta_rechazada/create_pregunta_rechazada.html'

    def get_success_url(self):
        return reverse('list_pregunta_rechazada')


class UpdatePreguntaRechazada(UpdateView):
    model = Pregunta_Rechazada
    fields = ['motivo', 'respuestaPersonalizada', 'comentarios',]
    template_name = 'pregunta_rechazada/edit_pregunta_rechazada.html'

    def get_success_url(self):
        return reverse('list_pregunta_rechazada')

class ListPreguntaRechazada(ListView):
    paginate_by = 50
    model = Pregunta_Rechazada
    fields = ['motivo', 'respuestaPersonalizada', 'comentarios',]
    template_name = 'pregunta_rechazada/list_pregunta_rechazada.html'

class DeletePreguntaRechazada(DeleteView):
    model = Pregunta_Rechazada
    success_url = reverse_lazy('list_pregunta_rechazada')

#### CRUD EMPRESA

class CreateEmpresa(CreateView):
    model = Empresa
    fields = ['nombre','sector','pais','website_corporativo','website_integridad','tot100','tot',]
    template_name = 'empresa/create_empresa.html'

    def get_success_url(self):
        return reverse('list_empresa')


class UpdateEmpresa(UpdateView):
    model = Empresa
    fields = ['nombre','sector','pais','website_corporativo','website_integridad','tot100','tot',]
    template_name = 'empresa/edit_empresa.html'

    def get_success_url(self):
        return reverse('list_empresa')

class ListEmpresa(ListView):
    paginate_by = 50
    model = Empresa
    fields = ['nombre','sector','pais','website_corporativo','website_integridad','tot100','tot',]
    template_name = 'empresa/list_empresa.html'

class DeleteEmpresa(DeleteView):
    model = Empresa
    success_url = reverse_lazy('list_empresa')

#### CRUD Articulo
class CreateArticulo(CreateView):
    model = Articulo
    fields = ['imagen','titulo','contenido','revista','autor','slug', 'url', 'fecha', ]
    template_name = 'articulo/create_articulo.html'

    def get_success_url(self):
        return reverse('list_articulo')


class UpdateArticulo(UpdateView):
    model = Articulo
    fields = ['imagen','titulo','contenido','revista','autor','slug', 'url', 'fecha', ]
    template_name = 'articulo/edit_articulo.html'

    def get_success_url(self):
        return reverse('list_articulo')


class ListArticulo(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    model = Articulo
    fields = ['imagen','titulo','contenido','revista','autor','slug', 'url', 'fecha', ]
    template_name = 'articulo/list_articulo.html'
    login_url = '/kondo-admin/'
    permission_required = 'is_admin'
    # redirect_field_name = 'redirect_to'

class DeleteArticulo(DeleteView):
    model = Articulo
    success_url = reverse_lazy('list_articulo')

#### CRUD Perfiles
class CreatePerfil(CreateView):
    model = Perfil
    fields = ['user','telefono_fijo','telefono_celular','empresa',]
    template_name = 'perfiles/create_perfiles.html'

    def get_success_url(self):
        return reverse('list_perfiles')


class UpdatePerfil(UpdateView):
    model = Perfil
    fields = ['user','telefono_fijo','telefono_celular','empresa',]
    template_name = 'perfiles/edit_perfiles.html'

    def get_success_url(self):
        return reverse('list_perfiles')

class ListPerfil(ListView):
    paginate_by = 10
    model = Perfil
    fields = ['user','telefono_fijo','telefono_celular','empresa',]
    template_name = 'perfiles/list_perfiles.html'

class DeletePerfil(DeleteView):
    model = Perfil
    success_url = reverse_lazy('list_perfiles')

#### CRUD GLOSARIO

class CreateGlosario(CreateView):
    model = Glosario
    fields = ['titulo','descripcion',]
    template_name = 'glosario/create_glosario.html'

    def get_success_url(self):
        return reverse('list_glosario')


class UpdateGlosario(UpdateView):
    model = Glosario
    fields = ['titulo','descripcion',]
    template_name = 'glosario/edit_glosario.html'

    def get_success_url(self):
        return reverse('list_glosario')

class ListGlosario(ListView):
    paginate_by = 15
    model = Glosario
    fields = ['titulo','descripcion',]
    template_name = 'glosario/list_glosario.html'

class DeleteGlosario(DeleteView):
    model = Glosario
    success_url = reverse_lazy('list_glosario')

#### END CRUD GLOSARIO


class CreateEntradasRecientes(CreateView):
    model = Entradas_Recientes
    fields = ['titulo','imagen', 'fecha', 'url']
    template_name = 'entradas_recientes/create_entradas_recientes.html'

    def get_success_url(self):
        return reverse('list_entradas_recientes')

class UpdateEntradasRecientes(UpdateView):
    model = Entradas_Recientes
    fields = ['titulo','imagen', 'fecha' , 'url']
    template_name = 'entradas_recientes/edit_entradas_recientes.html'

    def get_success_url(self):
        return reverse('list_entradas_recientes')

class ListEntradasRecientes(ListView):
    paginate_by = 10
    model = Entradas_Recientes
    fields = ['titulo','imagen', 'fecha', 'url']
    template_name = 'entradas_recientes/list_entradas_recientes.html'

class DeleteEntradasRecientes(DeleteView):
    model = Entradas_Recientes
    success_url = reverse_lazy('list_entradas_recientes')


#### CRUD FUENTES

class CreateFuente(CreateView):
    model = Fuentes
    fields = ['titulo','libro','url',]
    template_name = 'fuentes/create_fuente.html'

    def get_success_url(self):
        return reverse('list_fuente')

class UpdateFuente(UpdateView):
    model = Fuentes
    fields = ['titulo','libro','url',]
    template_name = 'fuentes/edit_fuente.html'

    def get_success_url(self):
        return reverse('list_fuente')

class ListFuente(ListView):
    paginate_by = 10
    model = Fuentes
    fields = ['titulo','libro','url',]
    template_name = 'fuentes/list_fuente.html'

class DeleteFuente(DeleteView):
    model = Fuentes
    success_url = reverse_lazy('list_fuente')

#### END CRUD FUENTES


#### CRUD PAISES

class CreatePaises(CreateView):
    model = Paises
    fields = ['pais',]
    template_name = 'pais/create_pais.html'

    def get_success_url(self):
        return reverse('list_paises')

class UpdatePaises(UpdateView):
    model = Paises
    fields = ['pais',]
    template_name = 'pais/edit_pais.html'

    def get_success_url(self):
        return reverse('list_paises')

class ListPaises(ListView):
    paginate_by = 10
    model = Paises
    fields = ['pais',]
    template_name = 'pais/list_pais.html'

class DeletePaises(DeleteView):
    model = Paises
    success_url = reverse_lazy('list_paises')

#### CRUD CUESTIONARIO
class CreateCuestionario(CreateView):
    model = Cuestionario
    fields = ['preguntas','Empresa',]
    template_name = 'cuestionario/create_cuestionario.html'

    def get_success_url(self):
        return reverse('list_cuestionario')

class UpdateCuestionario(UpdateView):
    model = Cuestionario
    fields = ['preguntas','Empresa',]
    template_name = 'cuestionario/edit_cuestionario.html'

    def get_success_url(self):
        return reverse('list_cuestionario')

class ListCuestionario(ListView):
    paginate_by = 50
    model = Cuestionario
    fields = ['preguntas','Empresa',]
    template_name = 'cuestionario/list_cuestionario.html'

class DeleteCuestionario(DeleteView):
    model = Cuestionario
    success_url = reverse_lazy('list_cuestionario')

#### END CRUD CUESTIONARIO

#### CRUD RESPUESTAS

class CreateRespuesta(CreateView):
    model = Respuestas
    fields = ['valor','opcion','catalogo_pregunta',]
    template_name = 'respuestas/create_respuestas.html'

    def get_success_url(self):
        return reverse('list_respuestas')

class UpdateRespuesta(UpdateView):
    model = Respuestas
    fields = ['valor','opcion','catalogo_pregunta',]
    template_name = 'respuestas/edit_respuestas.html'

    def get_success_url(self):
        return reverse('list_respuestas')

class ListRespuesta(ListView):
    paginate_by = 50
    model = Respuestas
    fields = ['valor','opcion','catalogo_pregunta',]
    template_name = 'respuestas/list_respuestas.html'

class DeleteRespuesta(DeleteView):
    model = Respuestas
    success_url = reverse_lazy('list_cuestionario')

#### END CRUD RESPUESTAS

#CRUD SECTOR
class CreateSector(CreateView):
    model = Sectores
    fields = ['nombre']
    template_name = 'sector/create_sector.html'

    def get_success_url(self):
        return reverse('list_sector')

class UpdateSector(UpdateView):
    model = Sectores
    fields = ['nombre']
    template_name = 'sector/edit_sector.html'

    def get_success_url(self):
        return reverse('list_sector')

class ListSector(ListView):
    paginate_by = 20
    model = Sectores
    fields = ['nombre']
    template_name = 'sector/list_sector.html'

class DeleteSector(DeleteView):
    model = Sectores
    success_url = reverse_lazy('list_sector')

##END CRUD SECTOR

#CRUD USER
class CreateUser(CreateView):
    model = User
    fields = ['username', 'first_name', 'last_name', 'email', 'password', 'groups', 'user_permissions', 'is_staff', 'is_active', 'is_superuser', 'last_login', 'date_joined',]
    template_name = 'user/create_user.html'

    def get_success_url(self):
        return reverse('list_user')

class UpdateUser(UpdateView):
    model = User
    fields = ['username', 'first_name', 'last_name', 'email', 'password', 'groups', 'user_permissions', 'is_staff', 'is_active', 'is_superuser', 'last_login', 'date_joined',]
    template_name = 'user/edit_user.html'

    def get_success_url(self):
        return reverse('list_user')

class ListUser(ListView):
    paginate_by = 20
    model = User
    fields = ['username', 'first_name', 'last_name', 'email', 'password', 'groups', 'user_permissions', 'is_staff', 'is_active', 'is_superuser', 'last_login', 'date_joined',]
    template_name = 'user/list_user.html'

class DeleteUser(DeleteView):
    model = User
    success_url = reverse_lazy('list_user')

#CRUD PREGUNTAS
class CreatePregunta(CreateView):
    model = Pregunta
    fields = ['reactivo', 'respuesta', 'status', 'comentarios',]
    template_name = 'pregunta/create_pregunta.html'

    def get_success_url(self):
        return reverse('list_pregunta')

class UpdatePregunta(UpdateView):
    model = Pregunta
    fields = ['reactivo', 'respuesta', 'status', 'comentarios',]
    template_name = 'pregunta/edit_pregunta.html'

    def get_success_url(self):
        return reverse('list_pregunta')

class ListPregunta(ListView):
    paginate_by = 50
    model = Pregunta
    fields = ['reactivo', 'respuesta', 'status', 'comentarios',]
    template_name = 'corte/list_corte.html'

class DeletePregunta(DeleteView):
    model = Pregunta
    success_url = reverse_lazy('list_corte')

#CRUD CORTE
class CreateCorte(CreateView):
    model = Corte
    fields = ['fecha_de_corte', 'aprovado']
    template_name = 'corte/create_corte.html'

    def get_success_url(self):
        return reverse('list_corte')

class UpdateCorte(UpdateView):
    model = Corte
    fields = ['fecha_de_corte', 'aprovado']
    template_name = 'corte/edit_corte.html'

    def get_success_url(self):
        return reverse('list_corte')

class ListCorte(ListView):
    paginate_by = 50
    model = Corte
    fields = ['fecha_de_corte', 'aprovado']
    template_name = 'corte/list_corte.html'

class DeleteCorte(DeleteView):
    model = Corte
    success_url = reverse_lazy('list_corte')


def index(request):
  return redirect('/login/')

def empresa(request):
  method = request.method
  user = request.user
  # user_group = user.groups.all()[0]
  if method == 'GET':
    if user.is_authenticated():
      corte = Corte.objects.all().order_by('-fecha_de_corte')
      corte = corte[0]
      usuario = Perfil.objects.get(user=user.pk)
      empresa = usuario.empresa
      cuestionario = Cuestionario.objects.get(Empresa=usuario.empresa.pk, Corte=corte.pk)

      preguntas = cuestionario.preguntas.all()
      preguntasCTX = {}
      for pregunta in preguntas:
          respuestas = Respuestas.objects.all().filter(catalogo_pregunta=pregunta.reactivo).values_list('opcion','pk')
          preguntasCTX[pregunta.reactivo.id_reactivo] = dict(respuestas)

      template = 'empresa/index.html'
      context = {
          'preguntas': preguntas,
          'preguntas_respuestas': json.dumps(preguntasCTX, cls=DjangoJSONEncoder),
          'empresa':empresa
      }
      return render(request, template, context)
    else:
      return redirect('/login/')
  else:
    pregunta_pk = request.POST.get('pregunta-pk')
    pregunta = Pregunta.objects.get(pk=pregunta_pk)
    respuesta_pk = request.POST.get('respuesta-pk')
    evidencia = request.POST.get('url')
    respuesta = Respuestas.objects.get(pk=respuesta_pk)
    respuesta.evidencia = evidencia
    pregunta.respuesta = respuesta
    pregunta.status = '0'
    pregunta.save()
    return redirect('/empresa/')


# def editInfo(request):
#   user = request.user
#   if user.is_authenticated:
#     empresa = Empresa.objects.get(nombre=user.first_name)
#     cuestionario = Cuestionario.objects.get(pk=empresa.cuestionario_id)
#     preguntas = cuestionario.preguntas.all()
#     print(preguntas[0].reactivo)

#     context = {
#       'empresa': empresa,
#       'preguntas': preguntas
#     }
#     return render(request, 'empresa/edit-info.html', context)
#   else:
#     return redirect(settings.LOGIN_URL)


def loginUser(request):
  method = request.method
  if method == 'POST':
    username = request.POST.get('username')
    password = request.POST.get('password')
    user = authenticate(request, username=username, password=password)
    if user is not None:
      group = user.groups.all()[0].name
      if group == 'empresa':
        login(request, user)
        return redirect(settings.LOGIN_REDIRECT_URL, user)
      elif group == 'revisor' :
        login(request, user)
        return redirect('/revisor/', user)
      else :
        login(request, user)
        # return redirect('/admin/', user)
        return redirect('/ic-admin/', user)
    else:
      return render(request, 'empresa/login.html', {'error': True})
  else:
    ctx = {}
    try:
      if request.GET['message']:
        ctx['message'] = '<p>Estimado usuario, Usted creó una cuenta en integridadcorporativa500.mx.</p>  <p>En breve recibira un correo con instrucciones.</p>'
    except:
      pass
    user = request.user
    if user.is_authenticated:
      group = user.groups.all()[0].name
      if group == 'empresa':
        login(request, user)
        return redirect(settings.LOGIN_REDIRECT_URL, user)
      elif group == 'revisor' :
        login(request, user)
        return redirect('/revisor/', user)
      else :
        login(request, user)
        return redirect('/ic-admin/', user)
    else:
      return render(request,'empresa/login.html', ctx )


def register(request):
    method = request.method
    if method == 'GET':
        user_form = SignUpForm
        perfil_form = PerfilForm()
        context = {
        'userForm': user_form,
        'perfilForm': perfil_form
        }
        return render(request, 'empresa/register.html', context)
    else:
        user_form = SignUpForm(request.POST)
        perfil_form = PerfilForm(request.POST)
        if user_form.is_valid() and perfil_form.is_valid():
            telefono_fijo = perfil_form.cleaned_data['telefono_fijo']
            telefono_celular = perfil_form.cleaned_data['telefono_celular']
            telefono_celular = perfil_form.cleaned_data['telefono_celular']
            empresa = perfil_form.cleaned_data['empresa']
            user = user_form.save()
            user.is_active = False
            user.save()
            group = Group.objects.get(name='empresa')
            group.user_set.add(user)
            group.save()
            profile = Perfil(user=user,telefono_fijo=telefono_fijo, telefono_celular=telefono_celular, empresa=empresa)
            profile.save()
            send_mail(
              'Solicitud de registro enviada | Contacto Integridad Corporativa',
              'Gracias por tu interés, en breve será contactado.',
              'contacto@integridadcorporativa500.mx',
              [user.email],
              fail_silently=False
            )
        else:
            print(user_form.errors)
            context = {
            'userForm': user_form,
            'perfilForm': perfil_form
            }
            return render(request, 'empresa/register.html', context)
        return redirect('/login/?message=true')

def rejectQuestion(request):
    if request.method == 'POST':
      pregunta_pk = request.POST.get('pregunta-pk')
      cuestionario_pk = request.POST.get('cuestionario-pk')
      empresa_pk = request.POST.get('empresa-pk')
      reject_form = Pregunta_Rechazada_Form(request.POST)
      if reject_form.is_valid():
        pregunta = Pregunta.objects.get(pk=pregunta_pk)
        form = reject_form.save()
        form.save()
        comentario_pk = form.pk
        comentario = Pregunta_Rechazada.objects.get(pk=comentario_pk)
        pregunta.comentarios = comentario
        pregunta.status = 2
        pregunta.save()
        return redirect('/validate/%s/%s/' %(cuestionario_pk, empresa_pk))

def modifyAnswer(request, pk):
  if request.method == 'GET' :
    pregunta = Pregunta.objects.get(pk=pk)
    pregunta = pregunta.__dict__
    respuesta = pregunta['respuesta']
    return HttpResponse(respuesta)
  else:
    question_id = request.POST.get('pregunta_id')
    pregunta = Pregunta.objects.get(pk=question_id)
    pregunta.respuesta = request.POST.get('respuesta')
    pregunta.save()
    return redirect('edit-info')


def articulos(request, slug):
  articulo = get_object_or_404(Articulo, slug=slug)
  template = 'micrositio/post.html'
  context = {
    'post': articulo
  }
  return render(request, template, context)

def validate(request, pk, empresa_pk):
  method = request.method
  if method == 'POST':
    pregunta_pk = request.POST.get('pregunta-pk')
    status = request.POST.get('status')
    if status == '1':
      pregunta = Pregunta.objects.get(pk=pregunta_pk)
      pregunta.status = status
      pregunta.save()
      return redirect('/validate/%s/%s/' %(pk, empresa_pk))
  else:
    template = 'revisor/validate.html'
    modalForm = Pregunta_Rechazada_Form()
    corte_anterior = Corte.objects.get(aprovado=True)
    cuestionario_anterior = Cuestionario.objects.get(Corte=corte_anterior, Empresa=empresa_pk)
    preguntas_anteriores = cuestionario_anterior.preguntas.all()
    cuestionario_actual = Cuestionario.objects.get(pk=pk)
    empresa = cuestionario_actual.Empresa
    preguntas_actuales = cuestionario_actual.preguntas.all()
    preguntas = zip(preguntas_anteriores, preguntas_actuales)
    context = {
      'preguntas': preguntas,
      'empresa': empresa,
      'form': modalForm,
      'pk': pk,
      'empresa_pk':empresa_pk
    }
    return render(request, template, context)

def glosario(request):
	if request.method == 'GET':
		data = serializers.serialize("json", Glosario.objects.all())
		return HttpResponse(data, content_type="application/json")

def fuentes(request):
  if request.method == 'GET':
    data = serializers.serialize("json", Fuentes.objects.all())
    return HttpResponse(data, content_type="application/json")

def entradasRecientes(request):
	if request.method == 'GET':
		data = serializers.serialize("json", Entradas_Recientes.objects.all().order_by('-fecha')[:3])
		return HttpResponse(data, content_type="application/json")

def blog_articulos(request):
	if request.method == 'GET':
		data = serializers.serialize("json", Articulo.objects.all())
		return HttpResponse(data, content_type="application/json")

def getArticleSlug(request,slug):
	if request.method == 'GET':
		article = serializers.serialize("json", Articulo.objects.filter(slug=slug))
		return HttpResponse(article, content_type="application/json")

def import_empresas(request):
  wb = load_workbook('mxvscorrupcion/501.xlsx')
  sheet = wb.get_sheet_by_name('Códigos')
  print(sheet)
  row1 = sheet.cell(row=3, column=3)
  print(row1)
  for line in range(9, 45):
    pregunta = Catalogo_Preguntas()
    pregunta.bloque = sheet.cell(row=line, column=1).value
    pregunta.id_reactivo = sheet.cell(row=line, column=2).value
    pregunta.descripcion = sheet.cell(row=line, column=3).value
    pregunta.save()
    respuestas = Respuestas()
    respuestas.opcion = sheet.cell(row=line, column=4).value
    if (respuestas.opcion is None):
      respuestas.opcion = 'N/A'
    respuestas.valor = '0'
    respuestas.catalogo_pregunta = pregunta
    respuestas.save()
    respuestas = Respuestas()
    respuestas.opcion = sheet.cell(row=line, column=5).value
    if (respuestas.opcion is None):
      respuestas.opcion = 'N/A'
    respuestas.valor = '0.5'
    respuestas.catalogo_pregunta = pregunta

    respuestas.save()
    respuestas = Respuestas()
    respuestas.opcion = sheet.cell(row=line, column=6).value
    if (respuestas.opcion is None):
      respuestas.opcion = 'N/A'
    respuestas.valor = '1'
    respuestas.catalogo_pregunta = pregunta

    respuestas.save()

  return HttpResponse(row1.value)

def getSector(sector):
  print('getSector ', sector)
  sector = Sectores.objects.filter(nombre=sector.strip())[0]
  return sector

def getCountry(country):
  country = Paises.objects.filter(pais=country)[0]
  return country

def getQuestion(qid):
  print('getQuestion', qid)
  if qid == 'IC_13C':
    qid = 'IC_13A'
  return Catalogo_Preguntas.objects.filter(id_reactivo=qid)[0]

def getAnswersByQid(qid):
  return Respuestas.objects.filter(catalogo_pregunta__pk=qid)


def import_empresas_final(request):
  wb = load_workbook('mxvscorrupcion/501.xlsx')
  sheet = wb.get_sheet_by_name('BD c valores COMP')
  print(sheet)
  row1 = sheet.cell(row=3, column=3)
  print(row1)
  # get questions ids
  # for line in range(9, 45):
  question_id_ary = []
  question_answer = []
  for question in range(7, 43):
    current_id = sheet.cell(row=1, column=question).value
    question_id_ary.append(current_id)

  # get client
  for client_Key in range(3, 503):
    client_data = {}
    client_answers = []
  # for client in range(3, 502):
    # for data in range(1, 6):

    name = sheet.cell(row=client_Key, column=2).value
    sector = sheet.cell(row=client_Key, column=3).value
    pais = sheet.cell(row=client_Key, column=4).value
    website_corporativo = sheet.cell(row=client_Key, column=5).value
    website_integridad = sheet.cell(row=client_Key, column=6).value

    print('name ', name)
    print('sector ', sector)
    client = Empresa()
    client.nombre = name
    client.sector = getSector(sector)
    client.pais = getCountry(pais)
    client.website_corporativo = website_corporativo
    client.website_integridad = website_integridad
    client.save()

    cuestionario = Cuestionario()
    print(dir(cuestionario))
    print(client.pk)

    for question in range(7, 43):
      current_answer = sheet.cell(row=client_Key, column=question).value
      question_answer.append(current_answer)
      qindex = question-7
      print(question_id_ary[qindex])



      question_id = question_id_ary[qindex]
      if question_id != 'IC_15A' and question_id != 'IC_16A' and question_id != 'IC_19A' and question_id != 'IC_22A' and question_id != 'IC_23A':
        question = getQuestion(question_id)
        answers = getAnswersByQid(question.pk)
        selected_answer = answers.filter(valor = current_answer)[0]
        # try:
        preguntaObj = Pregunta()
        preguntaObj.reactivo = question
        preguntaObj.respuesta = selected_answer
        preguntaObj.save()
        cuestionario.Empresa = client
        cuestionario.save()
        cuestionario.preguntas.add(preguntaObj)
        cuestionario.save()
        # except:
        #   print('''pregunta couldn't be added''')
        #   pass




      res = [question_id_ary, question_answer]


  return HttpResponse(res)

@csrf_exempt
def send_email(request):
  method = request.method
  if method == 'POST':
    name = request.POST.get('name')
    email = request.POST.get('email')
    message = request.POST.get('message')
    result = send_mail(
      'Haz recibido un correo de ' + name + ' | Contacto Integridad Corporativa',
      '%s\n\nEmail del remitente: %s' %(message, email),
      email,
      ['contacto@integridadcorporativa500.mx'],
      fail_silently=False
    )
    if result:
      return JsonResponse({'ok': True})
      

def usersAdmin(request):
  method =  request.method
  baseUrl = request.get_host()
  if method == 'POST':
    user_id = request.POST.get('user-id')
    user_email = request.POST.get('email')
    user = User.objects.get(pk=user_id)
    user.is_active = True
    user.save()
    result = send_mail(
        'Contacto Integridad Corporativa',
        'Estimado usuario,\n\nUsted creó una cuenta en integridadcorporativa500.mx\n\nHaga clic en el siguiente enlace para crear una contraseña e iniciar sesión: %s/password_reset/' %(baseUrl),
        'contacto@integridadcorporativa500.mx',
        [user_email],
        fail_silently=False
    )
    return redirect('/admin-users/')
  else:
    print(request.get_host())
    template = 'back_office/index.html'
    profiles = Perfil.objects.all()
    context = {
        'profiles': profiles
    }
    return render(request, template, context)

class Kondo_Admin(ListView):
  paginate_by = 5

  def get_queryset(self):
    return Corte.objects.all().order_by('-aprovado', 'fecha_de_corte')

class Corte_Detail(ListView):
  paginate_by = 50

  def get_queryset(self):
    self.corte = Corte.objects.get(pk=self.kwargs['pk'])
    return Cuestionario.objects.filter(Corte=self.corte)

def new_corte(request):
  method = request.method
  if method == 'POST':
    fecha = request.POST.get('fecha')
    corte = Corte(fecha_de_corte=fecha, aprovado=False)
    corte.save()
    cuestionarios = Cuestionario.objects.all()
    for cuestionario in cuestionarios:
      preguntas = cuestionario.preguntas.all()
      cuestionario.pk = None
      cuestionario.save()
      nuevo_cuestionario = cuestionario
      for pregunta in preguntas:
        reactivo = pregunta.reactivo
        respuesta = Respuestas.objects.get(pk=pregunta.respuesta.pk)
        catalogo_pregunta = respuesta.catalogo_pregunta
        respuesta.pk = None
        respuesta.save()
        nueva_respuesta = respuesta
        nueva_respuesta.catalogo_pregunta = catalogo_pregunta
        nueva_respuesta.save()
        pregunta.pk = None
        pregunta.save()
        nueva_pregunta = pregunta
        nueva_pregunta.respuesta = nueva_respuesta
        nueva_pregunta.reactivo = reactivo
        nueva_pregunta.save()
        nuevo_cuestionario.preguntas.add(nueva_pregunta)
      nuevo_cuestionario.Corte = corte
      nuevo_cuestionario.save()
  return redirect('kondo-admin')


def aprobar_corte(request):
  corte_pk = request.POST.get('corte-pk')
  corte_anterior = Corte.objects.get(aprovado=True)
  corte_anterior.aprovado = False
  corte_anterior.save()
  corte = Corte.objects.get(pk=corte_pk)
  corte.aprovado = True
  corte.save()
  return redirect('/revisor/')



def ic_admin(request):
  user = request.user
  if user.is_authenticated():
    template = 'ic_admin/index.html'
    return render(request, template)
  else:
    return redirect('/login/')










